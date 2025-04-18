"""
Excel Data Loader Module.

This module provides functionality to load data from Excel files (.xlsx, .xls, .xlsm)
into pandas DataFrames with support for multiple sheets and various Excel formats.
"""

import logging
from pathlib import Path
from typing import Optional, Dict, Any, Union, List, Tuple

import pandas as pd

from .base_loader import BaseDataLoader

# Configure logging
logger = logging.getLogger(__name__)


class ExcelDataLoader(BaseDataLoader):
    """
    Data loader for Excel files.
    
    This loader supports .xlsx, .xls, and .xlsm formats, with capabilities to read
    specific sheets, ranges, and handle various Excel features.
    """
    
    def __init__(self, verbose: bool = True):
        """
        Initialize the Excel data loader.
        
        Args:
            verbose: Whether to display progress bars and logging information.
        """
        super().__init__(verbose=verbose)
    
    def load(self, file_path: Union[str, Path], sheet_name: Union[str, int, List, None] = 0,
             header: Union[int, None] = 0, usecols: Optional[Union[List, str]] = None,
             skiprows: Optional[Union[List, int]] = None, 
             engine: Optional[str] = None, **kwargs) -> Union[pd.DataFrame, Dict[str, pd.DataFrame]]:
        """
        Load data from an Excel file into pandas DataFrames.
        
        Args:
            file_path: Path to the Excel file to load.
            sheet_name: Sheet(s) to read (0-indexed int for specific sheet, str for name, list for multiple, None for all).
            header: Row number to use as column names (0-indexed), or None if no header.
            usecols: Columns to read (str range like 'A:E' or list of indices).
            skiprows: Row indices to skip or number of rows to skip.
            engine: Excel engine to use ('openpyxl', 'xlrd', or None for auto-detection).
            **kwargs: Additional arguments passed to pandas.read_excel.
            
        Returns:
            A pandas DataFrame containing the loaded data if one sheet,
            or a dict of sheet_name -> DataFrame if multiple sheets.
            
        Raises:
            FileNotFoundError: If the file does not exist.
            PermissionError: If the file cannot be read due to permissions.
            ValueError: If the file format is invalid or unsupported.
        """
        # Validate the file path and check format
        file_path = self.validate_file(file_path)
        self._check_format(file_path)
        
        try:
            logger.info(f"Loading Excel file: {file_path}")
            
            # Determine engine if not specified
            if engine is None:
                engine = self._determine_engine(file_path)
            
            # Load the data
            result = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=header,
                usecols=usecols,
                skiprows=skiprows,
                engine=engine,
                **kwargs
            )
            
            # Log success info
            if isinstance(result, dict):
                logger.info(f"Successfully loaded {len(result)} sheets from {file_path}")
                for sheet, df in result.items():
                    logger.info(f"  - Sheet '{sheet}': {len(df)} rows, {len(df.columns)} columns")
            else:
                logger.info(f"Successfully loaded sheet with {len(result)} rows and {len(result.columns)} columns")
            
            return result
            
        except Exception as e:
            self.log_error(f"Error loading Excel file {file_path}", e)
            raise
    
    def load_specific_cells(self, file_path: Union[str, Path], 
                           sheet_name: Union[str, int] = 0,
                           start_cell: str = 'A1', 
                           end_cell: Optional[str] = None,
                           **kwargs) -> pd.DataFrame:
        """
        Load a specific range of cells from an Excel file.
        
        Args:
            file_path: Path to the Excel file.
            sheet_name: Sheet to read from.
            start_cell: The starting cell reference (e.g., 'A1').
            end_cell: The ending cell reference (e.g., 'C10'). If None, reads to the end.
            **kwargs: Additional arguments passed to pandas.read_excel.
            
        Returns:
            A pandas DataFrame containing the loaded data.
        """
        file_path = self.validate_file(file_path)
        
        try:
            # If end_cell is not specified, determine data range
            if end_cell is None:
                end_cell = self._find_data_range(file_path, sheet_name, start_cell)
            
            # Convert cell references to usable range for read_excel
            range_str = f"{start_cell}:{end_cell}"
            
            logger.info(f"Loading range {range_str} from sheet '{sheet_name}' in {file_path}")
            
            # Read the specified range
            return pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                usecols=range_str.split(':')[0][0] + ':' + range_str.split(':')[1][0],
                skiprows=int(range_str.split(':')[0][1:]) - 1 if range_str.split(':')[0][1:].isdigit() else 0,
                nrows=int(range_str.split(':')[1][1:]) - int(range_str.split(':')[0][1:]) + 1 if all(r.split(':')[0][1:].isdigit() and r.split(':')[1][1:].isdigit() for r in [range_str]) else None,
                **kwargs
            )
        
        except Exception as e:
            self.log_error(f"Error loading cell range {start_cell}:{end_cell} from {file_path}", e)
            raise
    
    def get_sheet_names(self, file_path: Union[str, Path]) -> List[str]:
        """
        Get the names of all sheets in an Excel file.
        
        Args:
            file_path: Path to the Excel file.
            
        Returns:
            A list of sheet names.
        """
        file_path = self.validate_file(file_path)
        
        try:
            engine = self._determine_engine(file_path)
            
            if engine == 'openpyxl':
                import openpyxl
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                sheet_names = workbook.sheetnames
                workbook.close()
                return sheet_names
            
            elif engine == 'xlrd':
                import xlrd
                workbook = xlrd.open_workbook(file_path, on_demand=True)
                sheet_names = workbook.sheet_names()
                workbook.release_resources()
                return sheet_names
            
            # Fallback using pandas
            with pd.ExcelFile(file_path, engine=engine) as excel_file:
                return excel_file.sheet_names
        
        except Exception as e:
            self.log_error(f"Error getting sheet names from {file_path}", e)
            raise
    
    def load_multiple_sheets(self, file_path: Union[str, Path], 
                            sheet_names: Optional[List[Union[str, int]]] = None, 
                            **kwargs) -> Dict[str, pd.DataFrame]:
        """
        Load multiple sheets from an Excel file.
        
        Args:
            file_path: Path to the Excel file.
            sheet_names: List of sheet names or indices to load. If None, loads all sheets.
            **kwargs: Additional arguments passed to pandas.read_excel.
            
        Returns:
            A dictionary mapping sheet names to pandas DataFrames.
        """
        file_path = self.validate_file(file_path)
        
        # If no sheet names provided, get all sheet names
        if sheet_names is None:
            sheet_names = self.get_sheet_names(file_path)
        
        logger.info(f"Loading {len(sheet_names)} sheets from {file_path}")
        
        return self.load(file_path, sheet_name=sheet_names, **kwargs)
    
    def _check_format(self, file_path: Path) -> None:
        """
        Check if the file is a valid Excel format.
        
        Args:
            file_path: Path to the file to check.
            
        Raises:
            ValueError: If the file is not a recognized Excel format.
        """
        valid_extensions = ['.xlsx', '.xls', '.xlsm', '.xlsb']
        if not any(file_path.name.lower().endswith(ext) for ext in valid_extensions):
            supported_formats = ", ".join(valid_extensions)
            raise ValueError(
                f"File {file_path} is not a recognized Excel format. "
                f"Supported formats: {supported_formats}"
            )
    
    def _determine_engine(self, file_path: Path) -> str:
        """
        Determine the appropriate Excel engine based on file extension.
        
        Args:
            file_path: Path to the Excel file.
            
        Returns:
            The Excel engine to use ('openpyxl' for newer formats, 'xlrd' for older).
        """
        # For .xls files, use xlrd
        if file_path.suffix.lower() == '.xls':
            return 'xlrd'
        
        # For other Excel formats, use openpyxl
        return 'openpyxl'
    
    def _find_data_range(self, file_path: Path, sheet_name: Union[str, int], 
                         start_cell: str) -> str:
        """
        Find the end cell of a data range starting from a given cell.
        
        Args:
            file_path: Path to the Excel file.
            sheet_name: The sheet to examine.
            start_cell: The starting cell reference.
            
        Returns:
            The end cell reference of the data range.
        """
        try:
            engine = self._determine_engine(file_path)
            
            if engine == 'openpyxl':
                import openpyxl
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                
                # Get the sheet
                if isinstance(sheet_name, int):
                    sheet = workbook.worksheets[sheet_name]
                else:
                    sheet = workbook[sheet_name]
                
                # Parse start cell
                col_letter = ''.join(c for c in start_cell if c.isalpha()).upper()
                row_num = int(''.join(c for c in start_cell if c.isdigit()))
                
                # Find the maximum non-empty row
                max_row = row_num
                for r in range(row_num, sheet.max_row + 1):
                    if all(sheet.cell(row=r, column=c).value is None 
                           for c in range(openpyxl.utils.column_index_from_string(col_letter), 
                                         sheet.max_column + 1)):
                        break
                    max_row = r
                
                # Find the maximum non-empty column
                max_col = openpyxl.utils.column_index_from_string(col_letter)
                for c in range(openpyxl.utils.column_index_from_string(col_letter), sheet.max_column + 1):
                    if all(sheet.cell(row=r, column=c).value is None 
                           for r in range(row_num, max_row + 1)):
                        break
                    max_col = c
                
                # Convert column index back to letter
                max_col_letter = openpyxl.utils.get_column_letter(max_col)
                
                workbook.close()
                return f"{max_col_letter}{max_row}"
            
            elif engine == 'xlrd':
                import xlrd
                workbook = xlrd.open_workbook(file_path, on_demand=True)
                
                # Get the sheet
                if isinstance(sheet_name, int):
                    sheet = workbook.sheet_by_index(sheet_name)
                else:
                    sheet = workbook.sheet_by_name(sheet_name)
                
                # Parse start cell
                col_letters = ''.join(c for c in start_cell if c.isalpha()).upper()
                row_num = int(''.join(c for c in start_cell if c.isdigit())) - 1
                col_num = self._excel_col_to_num(col_letters)
                
                # Find the maximum non-empty row and column
                max_row = row_num
                max_col = col_num
                
                for r in range(row_num, sheet.nrows):
                    for c in range(col_num, sheet.ncols):
                        if sheet.cell_value(r, c) not in ('', None):
                            max_row = max(max_row, r)
                            max_col = max(max_col, c)
                
                workbook.release_resources()
                
                # Convert column index back to letter and return result
                return f"{self._excel_num_to_col(max_col)}{max_row+1}"
            
            # Fallback - return a large range
            return "XFD1048576"  # Maximum Excel cell reference
        
        except Exception as e:
            self.log_error(f"Error finding data range in {file_path}", e)
            return "XFD1048576"  # Maximum Excel cell reference
    
    def _excel_col_to_num(self, col_str: str) -> int:
        """
        Convert Excel column letter to column number (0-indexed).
        
        Args:
            col_str: The column letter (e.g., 'A', 'BC').
            
        Returns:
            The 0-indexed column number.
        """
        num = 0
        for c in col_str:
            num = num * 26 + (ord(c.upper()) - ord('A') + 1)
        return num - 1
    
    def _excel_num_to_col(self, num: int) -> str:
        """
        Convert column number to Excel column letter.
        
        Args:
            num: The 0-indexed column number.
            
        Returns:
            The column letter (e.g., 'A', 'BC').
        """
        col_str = ""
        num += 1
        while num > 0:
            num, remainder = divmod(num - 1, 26)
            col_str = chr(ord('A') + remainder) + col_str
        return col_str 